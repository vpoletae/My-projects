# e-register card

# Import all needed modules
from tkinter import *
from PIL import Image, ImageTk
import openpyxl as opxl
from datetime import datetime
import smtplib

# Creating visuals
class Application(Frame):
    def __init__(self, master):
        super(Application, self).__init__(master)
        
        pil_image = Image.open("Cisco.png")
        pil_image = pil_image.resize((70, 34))
        self.image = ImageTk.PhotoImage(pil_image)
        
        self.mycolor_1 = '#%02x%02x%02x' % (21, 73, 94)
        self.mycolor_2 = '#%02x%02x%02x' % (70, 100, 120)
        self.mycolor_3 = '#%02x%02x%02x' % (139, 16, 1)
        self.mycolor_4 = '#%02x%02x%02x' % (43, 85, 147)
        self.mycolor_A = '#%02x%02x%02x' % (9, 125, 188)
        self.mycolor_T = '#%02x%02x%02x' % (0, 133, 22)
        self.grid()
        self.create_widgets()
        
    def create_widgets(self):

        # Label: Date
        Label (self,
               text = datetime.strftime(datetime.now(), "%d.%m.%Y"),
               font = ('Arial', 11, 'bold'),
               fg = 'black'
               ).grid(row = 0, column = 0, sticky = 'W')
        
        #Inserting Logo
        Label(self, image = self.image).grid(row = 0, column = 3, sticky = 'E')

        #Inserting names
        self.name = StringVar()
        self.name.set(None)
        
        names = ['R1','R2']
        column = 1
        colour = self.mycolor_A
        for name in names:
            Radiobutton(self,
                        text = name,
                        variable = self.name,
                        value = name,
                        font = ('Arial', 11, 'bold'),
                        fg = colour
                        ).grid(row = 0, column = column, sticky = 'W')
            column += 1
            colour = self.mycolor_T
            
        # Label: Sender
        Label (self,
               text = 'Отправитель',
               font = ('Arial', 10, 'bold'),
               fg = self.mycolor_1
               ).grid(row = 2, column = 0, sticky = 'W')
        self.sender_ent = Entry(self, bd = 2, fg = 'black')
        self.sender_ent.grid(row = 2, column = 1, columnspan = 3, sticky = 'NSEW')
        
        # Label: Recepient
        Label (self,
               text = 'Получатель',
               font = ('Arial', 10, 'bold'),
               fg = self.mycolor_1
               ).grid(row = 3, column = 0, sticky = 'W')
        self.recepient_ent = Entry(self, bd = 2, fg = 'black')
        self.recepient_ent.grid(row = 3, column = 1, columnspan = 3, sticky = 'NSEW')

         # Label: Doc_description
        Label (self,
               text = 'Наименование',
               font = ('Arial', 10, 'bold'),
               fg = self.mycolor_1
               ).grid(row = 4, column = 0, sticky = W)
        self.doc_description_ent = Entry(self, bd = 2, fg = 'black')
        self.doc_description_ent.grid(row = 4, column = 1, columnspan = 3, sticky = 'NSEW')

        # Label: -------------------
        Label (self,
               text = ' '*131,
               fg = self.mycolor_1
               ).grid(row = 5, column = 0, columnspan = 4, sticky = W)

        # Label: Doc_type
        Label (self,
               text = 'Тип',
               font = ('Arial', 10, 'bold'),
               fg = self.mycolor_1
               ).grid(row = 6, column = 0, sticky = W)

        self.doc_type = StringVar()
        self.doc_type.set(None)
        
        doc_types = ['Входящий','Исходящий','внутренний']
        column = 1
        for doc_type in doc_types:
            Radiobutton(self,
                        text = doc_type,
                        variable = self.doc_type,
                        value = doc_type,
                        ).grid(row = 6, column = column, sticky = 'W')
            column += 1

        # Label: -------------------
        Label (self,
               text = '-'*131,
               fg = self.mycolor_1
               ).grid(row = 7, column = 0, columnspan = 4, sticky = W)
 

        # Label: Doc_class
        Label (self,
               text = 'Класс',
               font = ('Arial', 10, 'bold'),
               fg = self.mycolor_1
               ).grid(row = 8, column = 0, rowspan = 3, sticky = 'W')

        self.doc_class = StringVar()
        self.doc_class.set(None)
        
        doc_classes = ['Контракт/ДС/Приложения',
                               'ТЗ/Спецификация',
                               'Доверенность',
                               'Счёт/Счёт-фактура/ТТН/Акт',
                               'Письмо/Уведомление/Запрос',
                               'Авансовый отчёт',
                               'HR документы',
                               'Маркетинг/Реклама/PR',
                               'Личное']
     
        column = 1
        row = 8
        interim_count = int()
        
        for doc_class in doc_classes:
            Radiobutton(self,
                        text = doc_class,
                        variable = self.doc_class,
                        value = doc_class
                        ).grid(row = row, column = column, sticky = 'W')

            interim_count += 1
            if interim_count % 3 == 0:
                column += 1
                row = 8
            else:
                row += 1
            
       # Label: -------------------
        Label (self,
               text = '-'*131,
               fg = self.mycolor_1
               ).grid(row = 11, column = 0, columnspan = 4, sticky = W)

        # Label: Courier_type
        Label (self,
               text = 'Вид доставки',
               font = ('Arial', 10, 'bold'),
               fg = self.mycolor_1
               ).grid(row = 12, column = 0, rowspan = 2, sticky = W)

        self.courier_type = StringVar()
        self.courier_type.set(None)
        
        courier_types = ['Почта России','Экспресс-доставка','Лично в руки', 'Факс', 'E-mail', 'Сотрудник']
        column = 1
        row = 12
        for courier_type in courier_types:
            Radiobutton(self,
                        text = courier_type,
                        variable = self.courier_type,
                        value = courier_type
                        ).grid(row = row, column = column, sticky = 'W')
            if column % 3 == 0:
                row += 1
                column = 0
            else:
                pass
            column += 1

        # Label: -------------------
        Label (self,
               text = '-'*131,
               fg = self.mycolor_1
               ).grid(row = 14, column = 0, columnspan = 4, sticky = W)

        # Label: Courier_type
        Label (self,
               text = 'Тип доставки',
               font = ('Arial', 10, 'bold'),
               fg = self.mycolor_1
               ).grid(row = 15, column = 0, sticky = W)

        self.courier_sort = StringVar()
        self.courier_sort.set(None)
        
        courier_sorts = ['Коробка', 'Пакет', 'Документы/Талоны']
        column = 1
        for courier_sort in courier_sorts:
            Radiobutton(self,
                        text = courier_sort,
                        variable = self.courier_sort,
                        value = courier_sort
                        ).grid(row = 15, column = column, sticky = 'W')
            column += 1

        # Label: -------------------
        Label (self,
               text = '-'*131,
               fg = self.mycolor_1
               ).grid(row = 16, column = 0, columnspan = 4, sticky = W)

        # Label: Comments
        Label (self,
               text = 'Комментарии',
               font = ('Arial', 10, 'bold'),
               fg = self.mycolor_1
               ).grid(row = 17, column = 0, sticky = W)
        self.comments_ent = Entry(self, bd = 2, fg = 'black')
        self.comments_ent.grid(row = 17, column = 1, columnspan = 3, sticky = 'NSEW')

        # Label: -------------------
        Label (self,
               text = ' '*131,
               fg = self.mycolor_1
               ).grid(row = 18, column = 0, columnspan = 4, sticky = W)
            
        # Button Register
        reg_button = Button(self,
                            text = 'Зарегистрировать документ',
                            command = self.register,
                            font = ('Arial', 12, 'bold'),
                            bg = self.mycolor_1,
                            fg = 'white')
        reg_button.config(height = 2, width = 15)
        reg_button.grid(row = 19, column = 0, columnspan = 4, sticky = 'NSEW')

        # Output field
        self.register_txt = Text(self, width = 82, height = 1, wrap = WORD, bd = 2, fg = 'black')
        self.register_txt.grid(row = 20, column = 0, columnspan = 4, pady = 1)

        # Label: -------------------
        Label (self,
               text = ' '*131,
               fg = self.mycolor_1
               ).grid(row = 21, column = 0, columnspan = 4, sticky = W)
        
        # Label: Reciever
        Label (self,
               text = 'Забрал сам',
               font = ('Arial', 10, 'bold'),
               fg = self.mycolor_1
               ).grid(row = 22, column = 0, sticky = W)
        self.reciever_ent = Entry(self, bd = 2, fg = 'black')
        self.reciever_ent.grid(row = 22, column = 1, columnspan = 2, sticky = 'NSEW')

        reg_button = Button(self,
                            text = 'Учесть',
                            command = self.check,
                            font = ('Arial', 10, 'bold'),
                            bg = self.mycolor_1,
                            fg = 'white')
        reg_button.config(height = 1)
        reg_button.grid(row = 22, column = 3, sticky = 'NSEW')
        
        # Button show1
        show1_button = Button(self,
                              text = '+',
                              command = self.show1,
                              font = ('Arial', 10, 'bold'),
                              bg = self.mycolor_1,
                              fg = 'white')
        show1_button.grid(row = 2, column = 5, sticky = 'W')

        # Button show2
        show2_button = Button(self,
                              text = '+',
                              command = self.show2,
                              font = ('Arial', 10, 'bold'),
                              bg = self.mycolor_1,
                              fg = 'white')
        show2_button.grid(row = 3, column = 5, sticky = 'W')

        # Button show3
        show3_button = Button(self,
                              text = '+',
                              command = self.show3,
                              font = ('Arial', 10, 'bold'),
                              bg = self.mycolor_1,
                              fg = 'white')
        show3_button.grid(row = 4, column = 5, sticky = 'W')
        
    def show1(self):
        lb=Listbox(self, width = 42)
        lb.grid(row = 1, column = 6, columnspan = 2, rowspan = 21, sticky = 'NSEW')

        wb = opxl.load_workbook('Reg_journal.xlsx')
        ws_sender = wb['Sender']
        col_A = ws_sender['A']
        sender_list = []
        for i, sender in enumerate(col_A):
            sender_list.append(sender.value)
        sender_list.sort(reverse = True)
        for sender in sender_list:
            lb.insert(0, sender)
    
        def choose():
            self.sender_ent.delete(0, END)
            self.sender_ent.insert(0, lb.get('active'))
            
        write_button = Button(self,
                              text = "\n".join("Заполнить"),
                              command = choose,
                              font = ('Arial', 9, 'bold'),
                              bg = self.mycolor_4,
                              fg = 'white')
        write_button.config(height = 29)
        write_button.grid(row =5, column = 5, rowspan = 18, sticky = 'NSEW')

        def add():
            wb = opxl.load_workbook('Reg_journal.xlsx')
            ws_sender = wb['Sender']
            sender = self.sender_ent.get()
            if not sender or sender == ' ':
                pass
            else:
                ws_sender.cell(column = 1, row = ws_sender.max_row +1).value = sender
            wb.save('Reg_journal.xlsx')
            
        def delete():
            wb = opxl.load_workbook('Reg_journal.xlsx')
            ws_sender = wb['Sender']
            col_A = ws_sender['A']
            to_del = lb.get('active')
            for i, send in enumerate(col_A):
                if send.value == to_del:
                    ws_sender.delete_rows(send.row, 1)
                    break
                else: continue
            wb.save('Reg_journal.xlsx')
        
        # Button add
        add_button = Button(self,
                              text = 'add',
                              command = add,
                              font = ('Arial', 10, 'bold'),
                              bg = self.mycolor_3,
                              fg = 'white')
        add_button.grid(row = 22, column = 6, sticky = 'NSEW')

        # Button Del
        add_button = Button(self,
                              text = 'delete',
                              command = delete,
                              font = ('Arial', 10, 'bold'),
                              bg = self.mycolor_4,
                              fg = 'white')
        add_button.grid(row = 22, column = 7, sticky = 'NSEW')
        
    def show2(self):
        lb=Listbox(self, width = 42)
        lb.grid(row = 1, column = 6, columnspan = 2, rowspan = 21, sticky = 'NSEW')

        wb = opxl.load_workbook('Reg_journal.xlsx')
        ws_recepient = wb['Recepient']
        col_A = ws_recepient['A']
        recepient_list = []
        for i, recepient in enumerate(col_A):
            recepient_list.append(recepient.value)
        recepient_list.sort(reverse = True)
        for recepient in recepient_list:
            lb.insert(0, recepient)

        def choose():
            self.recepient_ent.delete(0, END)
            self.recepient_ent.insert(0, lb.get('active'))
            
        write_button = Button(self,
                              text = "\n".join("Заполнить"),
                              command = choose,
                              font = ('Arial', 9, 'bold'),
                              bg = self.mycolor_4,
                              fg = 'white')
        write_button.config(height = 29)
        write_button.grid(row =5, column = 5, rowspan = 18, sticky = 'NSEW')

        def add():
            wb = opxl.load_workbook('Reg_journal.xlsx')
            ws_recepient = wb['Recepient']
            recepient = self.recepient_ent.get()
            if not recepient or recepient == ' ':
                pass
            else:
                ws_recepient.cell(column = 1, row = ws_recepient.max_row +1).value = recepient
            wb.save('Reg_journal.xlsx')
            
        def delete():
            wb = opxl.load_workbook('Reg_journal.xlsx')
            ws_recepient = wb['Recepient']
            col_A = ws_recepient['A']
            to_del = lb.get('active')
            for i, recepient in enumerate(col_A):
                if recepient.value == to_del:
                    ws_recepient.delete_rows(recepient.row, 1)
                    break
                else: continue
            wb.save('Reg_journal.xlsx')
        
        # Button add
        add_button = Button(self,
                              text = 'add',
                              command = add,
                              font = ('Arial', 10, 'bold'),
                              bg = self.mycolor_3,
                              fg = 'white')
        add_button.grid(row = 22, column = 6, sticky = 'NSEW')

        # Button Del
        add_button = Button(self,
                              text = 'delete',
                              command = delete,
                              font = ('Arial', 10, 'bold'),
                              bg = self.mycolor_4,
                              fg = 'white')
        add_button.grid(row = 22, column = 7, sticky = 'NSEW')
        
    def show3(self):
        lb=Listbox(self, width = 42) 
        lb.grid(row = 1, column = 6, columnspan = 2, rowspan = 21, sticky = 'NSEW')

        wb = opxl.load_workbook('Reg_journal.xlsx')
        ws_doc_type = wb['Doc_type']
        col_A = ws_doc_type['A']
        doc_type_list = []
        for i, doc_type in enumerate(col_A):
            doc_type_list.append(doc_type.value)
        doc_type_list.sort(reverse = True)
        for doc_type in doc_type_list:
            lb.insert(0, doc_type)

        def choose():
            self.doc_description_ent.delete(0, END)
            self.doc_description_ent.insert(0, lb.get('active'))
            
        write_button = Button(self,
                              text = "\n".join("Заполнить"),
                              command = choose,
                              font = ('Arial', 9, 'bold'),
                              bg = self.mycolor_4,
                              fg = 'white')
        write_button.config(height = 29)
        write_button.grid(row =5, column = 5, rowspan = 18, sticky = 'NSEW')

        def add():
            wb = opxl.load_workbook('Reg_journal.xlsx')
            ws_doc_type = wb['Doc_type']
            doc_type= self.doc_description_ent.get()
            if not doc_type or doc_type == ' ':
                pass
            else:
                ws_doc_type.cell(column = 1, row = ws_doc_type.max_row +1).value = doc_type
            wb.save('Reg_journal.xlsx')
            
        def delete():
            wb = opxl.load_workbook('Reg_journal.xlsx')
            ws_doc_type = wb['Doc_type']
            col_A = ws_doc_type['A']
            to_del = lb.get('active')
            for i, doc_type in enumerate(col_A):
                if doc_type.value == to_del:
                    ws_doc_type.delete_rows(doc_type.row, 1)
                    break
                else: continue
            wb.save('Reg_journal.xlsx')
        
        # Button add
        add_button = Button(self,
                              text = 'add',
                              command = add,
                              font = ('Arial', 10, 'bold'),
                              bg = self.mycolor_3,
                              fg = 'white')
        add_button.grid(row = 22, column = 6, sticky = 'NSEW')

        # Button Del
        add_button = Button(self,
                              text = 'delete',
                              command = delete,
                              font = ('Arial', 10, 'bold'),
                              bg = self.mycolor_4,
                              fg = 'white')
        add_button.grid(row = 22, column = 7, sticky = 'NSEW')
        
    # Key function
    def register(self):
        
        # get data from input
        current_date = datetime.strftime(datetime.now(), "%d.%m.%Y")
        name = self.name.get()
        sender = self.sender_ent.get().capitalize()
        recepient = self.recepient_ent.get().capitalize()
        doc_type = self.doc_type.get()
        doc_class = self.doc_class.get()
        courier = self.courier_type.get()
        courier_sort = self.courier_sort.get()
        doc_description = self.doc_description_ent.get()
        comments = self.comments_ent.get()
        

        # Manipulating with excel
        
        wb = opxl.load_workbook('Reg_journal.xlsx')
        ws_incoming = wb['Incoming']
        ws_outcoming = wb['Outcoming']
        ws_data = wb['Data']
        ws_contacts = wb['Contacts']
        col_A = ws_contacts['A']

        # Creating Reg.#
        
        reg = ''
        
        reg_contract_counter = int(ws_data.cell(row = 1, column = 2).value)
        reg_tz_counter = int(ws_data.cell(row = 2, column = 2).value)
        reg_power_of_attorney_counter = int(ws_data.cell(row = 3, column = 2).value)
        reg_accounting_counter = int(ws_data.cell(row = 4, column = 2).value)
        reg_letters_counter = int(ws_data.cell(row = 5, column = 2).value)
        reg_advance_reports_counter = int(ws_data.cell(row = 6, column = 2).value)
        reg_hr_counter = int(ws_data.cell(row = 7, column = 2).value)
        reg_personal_counter = int(ws_data.cell(row = 8, column = 2).value)
        reg_PR_counter = int(ws_data.cell(row = 9, column = 2).value)
        
        current_row_incoming = int(ws_data.cell(row = 10, column = 2).value)
        current_row_outcoming = int(ws_data.cell(row = 11, column = 2).value)

        # Changes in  reg # regarding to year (GAAP logic inside)
        current_year = 19 
        if datetime.strftime(datetime.now(), "%m") == '08':
            ws_data.cell(row = 1, column = 2).value = 0
            ws_data.cell(row = 2, column = 2).value = 0
            ws_data.cell(row = 3, column = 2).value = 0
            ws_data.cell(row = 4, column = 2).value = 0
            ws_data.cell(row = 5, column = 2).value = 0
            ws_data.cell(row = 6, column = 2).value = 0
            ws_data.cell(row = 7, column = 2).value = 0
            ws_data.cell(row = 8, column = 2).value = 0
            ws_data.cell(row = 9, column = 2).value = 0
            current_year += 1
        else:
            pass

        # Merging reg   
        if doc_class == 'Контракт/ДС/Приложения':
            reg += str(reg_contract_counter)
            reg_contract_counter += 1
            ws_data.cell(row = 1, column = 2).value = reg_contract_counter
        elif doc_class == 'ТЗ/Спецификация':
            reg += str(reg_tz_counter)
            reg_tz_counter += 1
            ws_data.cell(row = 2, column = 2).value = reg_tz_counter
        elif doc_class == 'Доверенность':
            reg += str(reg_power_of_attorney_counter)
            reg_power_of_attorney_counter += 1
            ws_data.cell(row = 3, column = 2).value = reg_power_of_attorney_counter
        elif doc_class == 'Счёт/Счёт-фактура/ТТН/Акт':
            reg += str(reg_accounting_counter)
            reg_accounting_counter += 1
            ws_data.cell(row = 4, column = 2).value = reg_accounting_counter
        elif doc_class == 'Письмо/Уведомление/Запрос':
            reg += str(reg_letters_counter)
            reg_letters_counter += 1
            ws_data.cell(row = 5, column = 2).value = reg_letters_counter
        elif doc_class == 'Авансовый отчёт':
            reg += str(reg_advance_reports_counter)
            reg_advance_reports_counter += 1
            ws_data.cell(row = 6, column = 2).value = reg_advance_reports_counter
        elif doc_class == 'HR документы':
            reg += str(reg_hr_counter)
            reg_hr_counter += 1
            ws_data.cell(row = 7, column = 2).value = reg_hr_counter
        elif doc_class == 'Личное':
            reg += str(reg_personal_counter)
            reg_personal_counter += 1
            ws_data.cell(row = 8, column = 2).value = reg_personal_counter
        elif doc_class == 'Маркетинг/Реклама/PR':
            reg += str(reg_PR_counter)
            reg_PR_counter += 1
            ws_data.cell(row = 9, column = 2).value = reg_PR_counter
            
        reg += '_'
        reg += doc_type[0]
        reg += doc_class[0]
        reg += '/'
        reg += str(current_year)
        reg += name[0]
            
        # Showing reg. number
        self.register_txt.delete(0.0, END)
        self.register_txt.insert(0.0, reg)

        # Fulfilling xls-table
        if doc_type == 'Входящий' or doc_type == 'Компания':
            ws_incoming.cell(row = current_row_incoming, column = 1).value = current_date
            ws_incoming.cell(row = current_row_incoming, column = 2).value = reg
            ws_incoming.cell(row = current_row_incoming, column = 3).value = sender
            ws_incoming.cell(row = current_row_incoming, column = 4).value = recepient
            ws_incoming.cell(row = current_row_incoming, column = 5).value = doc_description
            ws_incoming.cell(row = current_row_incoming, column = 6).value = doc_type
            ws_incoming.cell(row = current_row_incoming, column = 7).value = doc_class
            ws_incoming.cell(row = current_row_incoming, column = 8).value = courier
            ws_incoming.cell(row = current_row_incoming, column = 9).value = courier_sort
            ws_incoming.cell(row = current_row_incoming, column = 10).value = comments

            current_row_incoming +=1
            ws_data.cell(row = 10, column = 2).value = current_row_incoming
            
        elif doc_type == 'Исходящий':
            ws_outcoming.cell(row = current_row_outcoming, column = 1).value = current_date
            ws_outcoming.cell(row = current_row_outcoming, column = 2).value = reg
            ws_outcoming.cell(row = current_row_outcoming, column = 3).value = sender
            ws_outcoming.cell(row = current_row_outcoming, column = 4).value = recepient
            ws_outcoming.cell(row = current_row_outcoming, column = 5).value = doc_description
            ws_outcoming.cell(row = current_row_outcoming, column = 6).value = doc_type
            ws_outcoming.cell(row = current_row_outcoming, column = 7).value = doc_class
            ws_outcoming.cell(row = current_row_outcoming, column = 8).value = courier
            ws_outcoming.cell(row = current_row_outcoming, column = 9).value = courier_sort
            ws_outcoming.cell(row = current_row_outcoming, column = 10).value = comments

            current_row_outcoming +=1
            ws_data.cell(row = 11, column = 2).value = current_row_outcoming
        
        # Mailing block
        # Defining CONSTANTS and variables
        # Adding name to body
        name = ''
        for i in range(2, ws_contacts.max_row):
            if recepient == ws_contacts.cell(row = i, column = 1).value:
                name = ws_contacts.cell(row = i, column = 4).value
            else:
                continue

        MAIL = 'your_mail'
        
        SUBJECT_1 = '{} Вам письмо'.format(reg)
        TEXT_1 = '''{0}, добрый день!

Вам пришло письмо:

Документ: {1}
Отправитель: {2}
Регистрационный номер: {3}

В настоящее время документ находится на reception.
Вы можете забрать его самостоятельно, или после 17.00 он будет помещён в Ваш Tray

Спасибо!

P.S. По всем вопросам просьба обращаться на moscow-reception@cisco.com'''.format(name, doc_description, sender, reg)
        
        TEXT_2 = 'Документ зарегистрирован. Сообщение не отправлено!'

        SUBJECT_3 = '{} Статус Вашего исходящего отправления'.format(reg)
        TEXT_3 = '''Добрый день! 

Ваше письмо зарегистрировано и в ближайшее время будет отправлено

Регистрационный номер: {0}

Спасибо!'''.format(reg)
        
        message_1 = 'Subject: {}\n\n{}'.format(SUBJECT_1, TEXT_1).encode('UTF-8')
        message_2 = 'Subject: {}\n\n{}'.format(SUBJECT_3, TEXT_3).encode('UTF-8')        

        self.sender_ent.delete(0, END)
        self.recepient_ent.delete(0, END)
        self.doc_description_ent.delete(0, END)
        self.comments_ent.delete(0, END)
        
        # key loop
        for index, cell in enumerate(col_A):
            if cell.value == recepient:
                mail = ws_contacts.cell(row = cell.row, column = 2).value
            else:
                pass

        smtp_obj = smtplib.SMTP('your_data', 25)
        smtp_obj.starttls()
        try:
            if doc_type == 'Исходящий':
                try:
                    for index, cell in enumerate(col_A):
                        if cell.value == sender:
                            mail = ws_contacts.cell(row = cell.row, column = 2).value
                            smtp_obj.sendmail(MAIL, mail, message_2)
                            smtp_obj.quit()
                            break
                        else:
                            pass
                except:
                    self.register_txt.delete(0.0, END)
                    self.register_txt.insert(0.0, TEXT_2)            
            else:
                smtp_obj.sendmail(MAIL, mail, message_1)
                smtp_obj.quit()
        except:
            self.register_txt.delete(0.0, END)
            self.register_txt.insert(0.0, TEXT_2)

        wb.save('Reg_journal.xlsx')

    def check(self):

        wb = opxl.load_workbook('Reg_journal.xlsx')
        ws_incoming = wb['Incoming']

        reciever = self.reciever_ent.get().capitalize()
        for i in range(ws_incoming.max_row, ws_incoming.max_row - 20, -1):
            if reciever == ws_incoming.cell(row = i, column = 4).value:
                ws_incoming.cell(row = i, column = 11).value = 'Забрал сам'
            else:
                pass

        wb.save('Reg_journal.xlsx')
        
root = Tk()
root.title('Электронная карта регистрации')
app = Application(root)
root.mainloop()
