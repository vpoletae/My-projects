import openpyxl as opxl
from openpyxl import workbook
from datetime import datetime
import os

PATH = 'path'

def convert_to_xlsx(tenders_list, path=PATH):
      current_date = datetime.strftime(datetime.now(), "%d%m%Y")
      wb = opxl.Workbook()
      
      # creating 'Brief' tab
      ws_brief = wb.active
      ws_brief.title = 'Brief'
      
      names_list =['Тип контракта', 'Тема закупки', 'Дата начала', 'Дата окончания',
                   'Ссылка', 'Стоимость закупки (руб.)', 'Текущий статус', 'Фильтр',
                   'Место доставки', 'Условия доставки', 'Контактное лицо', 'Email',
                   'Факт. адрес', 'Город/Область', 'ИНН', 'Наим. организации','Тел.', 'Требования']
      
      name_counter = 1
      for name in names_list:
            ws_brief.cell(row = 1, column = name_counter).value = name
            name_counter += 1
            
      row_counter = 2
      for tender in tenders_list:
            col_counter = 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['contract_type']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['subject']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['start_date']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['end_date']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['purchase_link']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['purchase_price']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['current_status']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['filter_name']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['delivery_place']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['delivery_term']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['contact_person']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['email']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['fact_address']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['clean_address']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['inn']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['org_name']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['phone']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['requirements']
            row_counter += 1

      filter_name = tenders_list[0]['filter_name']
      file_name = f'{filter_name}_{current_date}.xlsx'
      wb.save(os.path.join(path, file_name))
      return file_name

def write_to_xlsx(file_name, tenders_list, path=PATH):
      wb = opxl.load_workbook(os.path.join(path, file_name))
      
      # creating 'Brief' tab
      ws_brief = wb['Brief']

      row_counter = ws_brief.max_row + 1
      for tender in tenders_list:
            col_counter = 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['contract_type']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['subject']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['start_date']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['end_date']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['purchase_link']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['purchase_price']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['current_status']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['filter_name']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['delivery_place']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['delivery_term']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['contact_person']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['email']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['fact_address']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['clean_address']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['inn']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['org_name']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['phone']
            col_counter += 1
            ws_brief.cell(row = row_counter,
                          column = col_counter).value = tender['requirements']
            row_counter += 1

      wb.save(os.path.join(path, file_name))
      
def get_files_in_folder(path=PATH):
      return os.listdir(path)





























      
      
