from lxml.html import fromstring
import requests
import itertools
import openpyxl as opxl
from datetime import datetime

URL = 'https://hh.ru/search/vacancy'
TEXT = 'Python'
AREA = 1
ORDER_BY = 'publication_time'
FIRST_X = 5
vacancies = []

class HH_scrapper():

      def make_request(self, page, url=URL, text=TEXT, area=AREA, order_by=ORDER_BY, params={}):
            params['text'] = text
            params['area'] = area
            params['order_by'] = order_by
            params['page'] = page
            response = requests.get(url, params=params, headers={'User-Agent':'api-test-agent'})
            return response

      def create_html_element(self, response):
            html_element = fromstring(response.text)
            return html_element

      def parse(self, html_element):
            vacancy = html_element.xpath("//div[@class='resume-search-item__name']//a/text()")
            reference = html_element.xpath("//div[@class='resume-search-item__name']/a/@href")
            company = html_element.xpath("//div[@class='vacancy-serp-item__meta-info']//a/text()")
            desciption = html_element.xpath("//div[@class='vacancy-serp-item__info']/div \
                                                            [@data-qa='vacancy-serp__vacancy_snippet_responsibility']/text()")
            salary = html_element.xpath("//div[@class='vacancy-serp-item__compensation']/text()")
            salary = [i.replace('\xa0', '') for i in salary]
            date = html_element.xpath("//span[@class='vacancy-serp-item__publication-date']//text()")
            date = [i.replace('\xa0', ' ') for i in date]
            for i in itertools.zip_longest(vacancy, reference, company, desciption, salary, date):
                  vacancies.append(i)
            return vacancies

      def write_to_xlsx(self, vacancies):
            current_date = datetime.strftime(datetime.now(), "%d%m%Y")
            wb = opxl.Workbook()
            ws = wb.active
            
            names_list =['Vacancy', 'Reference', 'Company', 'Desciption', 'Salary', 'Date']
            name_counter = 1
            for name in names_list:
                  ws.cell(row = 1, column = name_counter).value = name
                  name_counter += 1

            row_counter = 2
            for vacancy in vacancies:
                  col_counter = 1
                  try:
                        ws.cell(row = row_counter, column = col_counter).value = vacancy[0].encode("utf8")
                        col_counter += 1
                        ws.cell(row = row_counter, column = col_counter).value = vacancy[1].encode("utf8")
                        col_counter += 1
                        ws.cell(row = row_counter, column = col_counter).value = vacancy[2].encode("utf8")
                        col_counter += 1
                        ws.cell(row = row_counter, column = col_counter).value = vacancy[3].encode("utf8")
                        col_counter += 1
                        ws.cell(row = row_counter, column = col_counter).value = vacancy[4].encode("utf8")
                        col_counter += 1
                        ws.cell(row = row_counter, column = col_counter).value = vacancy[5].encode("utf8")
                        row_counter += 1
                  except AttributeError:
                        pass
            wb.save(f'Python_vacancies_Moscow_{current_date}.xlsx')

scrapper = HH_scrapper()
for page in range(1,FIRST_X):
      response = scrapper.make_request(page)
      html_element = scrapper.create_html_element(response)
      vacancies = scrapper.parse(html_element)
scrapper.write_to_xlsx(vacancies)
        

