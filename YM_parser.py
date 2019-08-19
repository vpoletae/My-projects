import json
import requests
import openpyxl as opxl
from openpyxl import workbook
import time

class YandexMarketContent(object):
      API_VERSION = '2.1'

      def __init__(self):
            """init new object with key and token"""
            self.key = 'Your_key'
            self.token = 'Your_token'

      def make_url(self, resource, version=API_VERSION):
            """create first part of URL string"""
            return 'https://api.content.market.yandex.ru/v{0}/{1}'.format(version, resource)

      def make_request(self, resource, version=API_VERSION, params={}):
            """send response to web-server"""
            url = self.make_url(resource, version) 
            params['Authorization'] = self.key
            params['X-User-Authorization'] = self.token

            response = requests.get(url, params=params,
                                    headers={'Authorization':self.key, 'X-User-Authorization':self.token})
            output = json.loads(response.content.decode("utf-8"))
            if response.status_code == 401:
                  server_response = []
                  for error in output['errors']:
                        server_response.append(error)
                  raise YandexMarketContent.NotAuthorized(
                        '''Your key {0}, token {1} were not authorized at Yandex.Market API.
                        Server response: {2}'''.format(self.key, self.token, server_response))
            return output

      # special methods for parsing
      def show_high_level_regions(self):
            """
            returns top-level regions
            returns list of tuples (id, country)
            """
            response = self.make_request('geo/regions')
            regions = response['regions']

            top_level_id_list = []
            for region in regions:
                  top_level_id_list.append((region['id'], region['name']))
            return top_level_id_list
      
      def show_regions(self, id_list):
            """
            return ids of current districts
            returns list of tuples (id, concatenate)
            """
            children_regions_list = []
            for id_ in id_list:
                  response = self.make_request('geo/regions/{}/children'.format(id_[0]) )
                  regions = response['regions']
                  for region in regions:
                        children_regions_list.append((region['id'], region['name'] + ' ' + id_[1]))
            return children_regions_list
      
      def search_models(self, brand, geo_id):
            """
            returns stores regarding to area
            returns list of dict {'id':..., 'name':...}
            """
            response = self.make_request('search/filters',
                                         params={'geo_id':geo_id, 'text':brand})
            shop_dict_list = response['filters']
            clean_shops = []
            for shop_dict in shop_dict_list:
                  if shop_dict['id'] == '-6':
                        clean_shops = shop_dict['values']
                  else:
                        pass
            return clean_shops

      def find_shop_details_by_id(self, id_, fields = 'ALL'):
            """
            returns info about shop
            returns dict of dicts
            """
            response = self.make_request('shops/{0}'.format(id_), params={'fields':fields})

            shop_output = tuple()
            shop_description = {}
            
            shop_brand_name = response['shop']['name']
            shop_id = response['shop']['id']
            shop_domain = response['shop']['domain']
            shop_address = ''#response['shop']['organizations'][0]['address']
            shop_url = response['shop']['organizations'][0]['contactUrl']
            shop_name = response['shop']['organizations'][0]['name']
            shop_ogrn = response['shop']['organizations'][0]['ogrn']
            shop_postal_address = response['shop']['organizations'][0]['postalAddress']
            shop_type = response['shop']['organizations'][0]['type']
            shop_outlets = len(response['shop']['outlets'])
            shop_rating = response['shop']['rating']['count']
            shop_registered = response['shop']['registered']

            shop_description['shop_id'] = shop_id
            shop_description['shop_domain'] = shop_domain
            shop_description['shop_address'] = shop_address
            shop_description['shop_url'] = shop_url
            shop_description['shop_name'] = shop_name
            shop_description['shop_ogrn'] = shop_ogrn
            shop_description['shop_postal_address'] = shop_postal_address
            shop_description['shop_type'] = shop_type
            shop_description['shop_outlets'] = shop_outlets
            shop_description['shop_rating'] = shop_rating
            shop_description['shop_registered'] = shop_registered

            shop_output = (shop_brand_name, shop_description)
            
            return shop_output

      def find_shop_details(self, id_, fields = 'ALL'):
            response = self.make_request('shops/{0}'.format(id_), params={'fields':fields})
            pprint (response)
            
def parse(brand, list_of_store_keys):
      """
      returns full shop info
      dict-dict-list-dict
      """
      detailed_shop_list = []
      mistakes_list = []
      store_info = ''
      for store_key in list_of_store_keys:
            try:
                  store_info = api.find_shop_details_by_id(store_key)
                  detailed_shop_list.append(store_info)
                  print('...')
                  time.sleep(2)
            except KeyError:
                  mistakes_list.append(store_key)

      return detailed_shop_list, mistakes_list

def convert_to_xlsx(detailed_shop_list):
      
      wb = opxl.Workbook()
      ws = wb.active
      
      # creating top-line
      names_list =['brand_name', 'shop_id', 'shop_domain', 'shop_address', 'shop_url',
                        'shop_name', 'shop_ogrn', 'shop_postal_address',
                        'shop_type', 'shop_outlets', 'shop_rating', 'shop_registered']
      name_counter = 1
      for name in names_list:
            ws.cell(row = 1, column = name_counter).value = name
            name_counter += 1

      # writing down data into xlsx
      row_counter = 2
      for shop in detailed_shop_list:
            col_counter = 1
            ws.cell(row = row_counter, column = col_counter).value = shop[0]
            col_counter += 1
            ws.cell(row = row_counter, column = col_counter).value = shop[1]['shop_id']
            col_counter += 1
            ws.cell(row = row_counter, column = col_counter).value = shop[1]['shop_domain']
            col_counter += 1
            ws.cell(row = row_counter, column = col_counter).value = shop[1]['shop_address']
            col_counter += 1
            ws.cell(row = row_counter, column = col_counter).value = shop[1]['shop_url']
            col_counter += 1
            ws.cell(row = row_counter, column = col_counter).value = shop[1]['shop_name']
            col_counter += 1
            ws.cell(row = row_counter, column = col_counter).value = shop[1]['shop_ogrn']
            col_counter += 1
            ws.cell(row = row_counter, column = col_counter).value = shop[1]['shop_postal_address']
            col_counter += 1
            ws.cell(row = row_counter, column = col_counter).value = shop[1]['shop_type']
            col_counter += 1
            ws.cell(row = row_counter, column = col_counter).value = shop[1]['shop_outlets']
            col_counter += 1
            ws.cell(row = row_counter, column = col_counter).value = shop[1]['shop_rating']
            col_counter += 1
            ws.cell(row = row_counter, column = col_counter).value = shop[1]['shop_registered']
            row_counter += 1
                  
      wb.save('output.xlsx')
      
if __name__=='__main__':
      print('''
<Stores parser>

- search & parse store data across CIS
- feel free to enter any request
============================
            ''')
      while True:
            start_time = time.time()
            api = YandexMarketContent()
            brand = input('Enter your request: ')
            top_level_id_list = api.show_high_level_regions()
            districts_list = api.show_regions(top_level_id_list)
            regions_list = api.show_regions(districts_list)

            list_of_stores = []
            for region in regions_list:
                  list_of_stores.extend(api.search_models(brand, region[0]))
                  print('...')
                  time.sleep(1)

            set_of_keys_stores = set()
            for store in list_of_stores:
                  set_of_keys_stores.add(store['id'])

            list_of_store_keys = list(set_of_keys_stores)
            
            detailed_shop_list, mistakes_list = parse(brand, list_of_store_keys)
            convert_to_xlsx(detailed_shop_list)

            print('Stores found: {0}'.format(len(detailed_shop_list)))
            print('Store ids not identified:')
            print(mistakes_list)
            print("--- %s seconds ---" % (time.time() - start_time))
            print()
      




























































































      
      
